from __future__ import annotations

import asyncio
import io
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from fastapi.testclient import TestClient

from slice_helper.app import create_app
from slice_helper.config import Settings
from slice_helper.database import Database
from slice_helper.excel_export import HEADERS, TEMPLATE_PATH
from slice_helper.media import MediaError
from slice_helper.models import MediaProbe
from slice_helper.orchestrator import Orchestrator
from slice_helper.source_download import SourceDownloadError
from slice_helper.time_ocr import OcrText, TimeReference


def make_settings(tmp_path: Path) -> Settings:
    return Settings(
        islice_base_url="http://islice.test",
        public_base_url="http://helper.test",
        host="127.0.0.1",
        port=8090,
        data_dir=tmp_path / "data",
        temp_dir=tmp_path / "temp",
        ffmpeg_path="ffmpeg",
        ffprobe_path="ffprobe",
        max_active_jobs=1,
        poll_interval_seconds=0.01,
        window_timeout_seconds=1,
        ffmpeg_timeout_seconds=10,
    )


def create_channel(client: TestClient, name: str = "测试频道") -> str:
    response = client.post("/api/channels", json={"name": name})
    assert response.status_code == 201
    return response.json()["id"]


def test_job_api_control_and_database_backed_chunk_route(
    tmp_path: Path, monkeypatch
) -> None:
    resplit_call: dict[str, object] = {}
    overlap_call: dict[str, object] = {}

    async def fake_probe(_self, _path):
        return MediaProbe(
            duration=120.0,
            format_name="mpegts",
            video_codec="h264",
            audio_codec="aac",
        )

    async def fake_time_reference(_self, _source, frame_path, **_kwargs):
        frame_path.parent.mkdir(parents=True, exist_ok=True)
        frame_path.write_bytes(b"png")
        timestamp = datetime.fromisoformat("2026-06-20T12:29:59")
        return TimeReference(
            source_start_time=timestamp,
            observed_time=timestamp,
            frame_offset_seconds=0.0,
            matched_text="2026-06-20 12:29:59",
            confidence=0.99,
            ocr_texts=(OcrText("2026-06-20 12:29:59", 0.99),),
        )

    async def no_start(_self):
        return None

    async def no_stop(_self):
        return None

    async def fake_resplit(_self, job_id, window_index, task_id):
        resplit_call.update(
            job_id=job_id, window_index=window_index, task_id=task_id
        )
        return {
            "jobId": job_id,
            "windowIndex": window_index,
            "taskId": task_id,
            "status": "resplit_queued",
        }

    async def fake_accept_overlap(_self, job_id, window_index, task_id):
        overlap_call.update(
            job_id=job_id, window_index=window_index, task_id=task_id
        )
        return {
            "jobId": job_id,
            "windowIndex": window_index,
            "taskId": task_id,
            "status": "overlap_accepted",
        }

    monkeypatch.setattr("slice_helper.media.MediaService.probe", fake_probe)
    monkeypatch.setattr(
        "slice_helper.media.MediaService.detect_time_reference", fake_time_reference
    )
    monkeypatch.setattr(Orchestrator, "start", no_start)
    monkeypatch.setattr(Orchestrator, "stop", no_stop)
    monkeypatch.setattr(Orchestrator, "schedule_resplit", fake_resplit)
    monkeypatch.setattr(
        Orchestrator, "accept_resplit_overlap", fake_accept_overlap
    )

    source = tmp_path / "source.ts"
    source.write_bytes(b"source")
    configured = make_settings(tmp_path)

    with TestClient(create_app(configured)) as client:
        home = client.get("/")
        assert home.status_code == 200
        assert 'id="segmentPreview"' in home.text
        assert 'id="segmentPageInfo"' in home.text
        assert 'id="previousSegmentPage"' in home.text
        assert 'id="nextSegmentPage"' in home.text
        assert 'id="previewDialog"' not in home.text
        assert 'class="media-workspace"' in home.text
        assert 'class="segment-results-panel"' in home.text
        assert 'class="window-timeline-panel"' in home.text
        assert 'id="resplitDialog"' in home.text
        assert 'id="resplitForm"' in home.text
        assert 'id="segmentEditDialog"' in home.text
        assert 'id="segmentEditForm"' in home.text
        assert 'id="restoreSegmentEditButton"' in home.text
        assert 'id="segmentDetailDialog"' in home.text
        assert 'id="segmentExportDetailGrid"' in home.text
        assert "Excel 字段" in home.text
        assert "对应系统字段" in home.text
        assert "本条导出值" in home.text
        assert "确认重新拆分" in home.text
        assert "<th>操作</th>" in home.text
        assert "小任务进度" in home.text
        assert "下发时间" in home.text
        assert "<th>源偏移</th>" not in home.text
        assert "<th>窗口</th>" in home.text
        assert "<th>节目类型</th>" in home.text
        assert "<th>新闻事件</th>" in home.text
        assert "已审核" in home.text
        for content_type in (
            "新闻", "电视剧", "电影", "综艺", "少儿", "体育", "纪录片",
            "科教", "文艺", "生活服务", "商业广告", "公益广告", "电视购物", "其他",
        ):
            assert f'<option value="{content_type}">{content_type}</option>' in home.text
        assert 'id="summaryISlice"' in home.text
        assert "/static/styles.css?v=0.12.2" in home.text
        assert "/static/app.js?v=0.12.2" in home.text
        assert 'id="tailRebuildDialog"' in home.text
        assert 'id="tailRebuildConfirmation"' in home.text
        assert "TS 路径或 HTTP 地址" in home.text
        assert 'id="manageChannelsButton"' in home.text
        assert 'id="jobPageInfo"' in home.text
        app_js = client.get("/static/app.js")
        styles_css = client.get("/static/styles.css")
        assert app_js.status_code == 200
        assert styles_css.status_code == 200
        assert "从此窗口重跑" in app_js.text
        assert "旧 iSlice 任务和临时文件不会删除" in home.text
        assert "is-latest-submitted" in app_js.text
        assert "is-latest-submitted" in styles_css.text
        assert "width: 48px" in styles_css.text

        channel_id = create_channel(client)

        created = client.post(
            "/api/jobs",
            json={
                "sourcePath": str(source.resolve()),
                "channelId": channel_id,
                "broadcastDate": "2026-06-20",
                "templateId": "general",
                "language": "zh",
                "cutMode": "copy",
            },
        )
        assert created.status_code == 201
        job = created.json()
        job_id = job["id"]
        assert job["total_windows"] == 1
        assert job["status"] == "pending_schedule"
        assert job["islice_base_url"] == ""
        assert job["channel_name"] == "测试频道"
        assert job["reviewed"] is False

        reviewed = client.patch(f"/api/jobs/{job_id}/review", json={"reviewed": True})
        assert reviewed.status_code == 200
        assert reviewed.json()["reviewed"] is True
        assert client.get(f"/api/jobs/{job_id}").json()["job"]["reviewed"] is True
        assert client.get(f"/api/jobs/{job_id}/result").json()["job"]["reviewed"] is True
        assert client.patch("/api/jobs/missing/review", json={"reviewed": True}).status_code == 404
        assert job["broadcast_date"] == "2026-06-20"
        assert job["program_start_time"] == "2026-06-20T12:29:59"
        assert job["time_reference_source"] == "ocr"
        assert job["time_reference_text"] == "2026-06-20 12:29:59"
        assert job["time_reference_confidence"] == 0.99
        assert Path(job["time_reference_frame_path"]).is_file()

        listed = client.get("/api/jobs").json()
        assert listed["total"] == 1
        assert [item["id"] for item in listed["items"]] == [job_id]

        paused = client.post(f"/api/jobs/{job_id}/pause")
        assert paused.status_code == 200
        assert paused.json()["status"] == "paused"
        resumed = client.post(f"/api/jobs/{job_id}/resume")
        assert resumed.status_code == 200
        assert resumed.json()["status"] == "pending_schedule"

        chunk = configured.temp_dir / job_id / "window-000.ts"
        chunk.parent.mkdir(parents=True)
        chunk.write_bytes(b"test-ts-body")

        async def record_chunk() -> None:
            database = Database(configured.database_path)
            window = await database.upsert_window(job_id, 0, 0, 120)
            await database.update_window(
                window["id"], chunk_path=str(chunk), chunk_url="http://helper/chunk"
            )

        asyncio.run(record_chunk())
        response = client.get(f"/internal/chunks/{job_id}/0.ts")
        assert response.status_code == 200
        assert response.content == b"test-ts-body"
        assert response.headers["content-type"].startswith("video/mp2t")
        assert response.headers["content-length"] == str(len(b"test-ts-body"))
        assert client.get("/internal/chunks/not-a-job/0.ts").status_code == 404

        resplit = client.post(
            f"/api/jobs/{job_id}/windows/0/resplit",
            json={"taskId": "sh-test-w000-a1"},
        )
        assert resplit.status_code == 202
        assert resplit.json()["taskId"] == "sh-test-w000-a1"
        assert resplit_call == {
            "job_id": job_id,
            "window_index": 0,
            "task_id": "sh-test-w000-a1",
        }

        overlap = client.post(
            f"/api/jobs/{job_id}/windows/0/accept-overlap",
            json={"taskId": "sh-test-w000-a1"},
        )
        assert overlap.status_code == 200
        assert overlap.json()["status"] == "overlap_accepted"
        assert overlap_call == {
            "job_id": job_id,
            "window_index": 0,
            "task_id": "sh-test-w000-a1",
        }

        stopped = client.post(f"/api/jobs/{job_id}/stop")
        assert stopped.status_code == 200
        assert stopped.json()["status"] == "stopped"


def test_create_job_rejects_non_absolute_ts_path(tmp_path: Path) -> None:
    configured = make_settings(tmp_path)
    with TestClient(create_app(configured)) as client:
        response = client.post(
            "/api/jobs",
            json={"sourcePath": "relative.ts", "templateId": "general", "language": "zh"},
        )
    assert response.status_code == 422


def test_create_job_downloads_http_source_to_managed_path(
    tmp_path: Path, monkeypatch
) -> None:
    downloaded: dict[str, object] = {}

    async def fake_download(_self, url, target):
        downloaded["url"] = url
        downloaded["target"] = target
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(b"downloaded-ts")
        return len(b"downloaded-ts")

    async def fake_probe(_self, path):
        assert path == downloaded["target"]
        return MediaProbe(
            duration=7200.0,
            format_name="mpegts",
            video_codec="h264",
            audio_codec="aac",
        )

    async def failed_time_reference(_self, _source, _frame_path, **_kwargs):
        raise MediaError("timestamp not found")

    async def no_start(_self):
        return None

    async def no_stop(_self):
        return None

    monkeypatch.setattr(
        "slice_helper.source_download.HttpSourceDownloader.download", fake_download
    )
    monkeypatch.setattr("slice_helper.media.MediaService.probe", fake_probe)
    monkeypatch.setattr(
        "slice_helper.media.MediaService.detect_time_reference", failed_time_reference
    )
    monkeypatch.setattr(Orchestrator, "start", no_start)
    monkeypatch.setattr(Orchestrator, "stop", no_stop)

    configured = make_settings(tmp_path)
    source_url = "https://media.test/archive/day.ts?token=internal"
    with TestClient(create_app(configured)) as client:
        channel_id = create_channel(client)
        response = client.post(
            "/api/jobs",
            json={
                "sourcePath": source_url,
                "channelId": channel_id,
                "broadcastDate": "2026-06-20",
            },
        )

    assert response.status_code == 201
    job = response.json()
    managed_source = configured.data_dir / "jobs" / job["id"] / "source.ts"
    assert downloaded == {"url": source_url, "target": managed_source}
    assert managed_source.read_bytes() == b"downloaded-ts"
    assert job["source_url"] == source_url
    assert job["source_path"] == str(managed_source)
    assert job["source_size"] == len(b"downloaded-ts")
    assert job["total_windows"] == 2


def test_create_job_cleans_failed_http_download(tmp_path: Path, monkeypatch) -> None:
    async def failed_download(_self, _url, target):
        target.parent.mkdir(parents=True, exist_ok=True)
        target.with_name("source.partial.ts").write_bytes(b"partial")
        raise SourceDownloadError("remote connection closed")

    async def no_start(_self):
        return None

    async def no_stop(_self):
        return None

    monkeypatch.setattr(
        "slice_helper.source_download.HttpSourceDownloader.download", failed_download
    )
    monkeypatch.setattr(Orchestrator, "start", no_start)
    monkeypatch.setattr(Orchestrator, "stop", no_stop)
    configured = make_settings(tmp_path)

    with TestClient(create_app(configured)) as client:
        channel_id = create_channel(client)
        response = client.post(
            "/api/jobs",
            json={
                "sourcePath": "http://media.test/broken.ts",
                "channelId": channel_id,
                "broadcastDate": "2026-06-20",
            },
        )

    assert response.status_code == 502
    assert response.json()["detail"] == "remote connection closed"
    jobs_dir = configured.data_dir / "jobs"
    assert not jobs_dir.exists() or not list(jobs_dir.iterdir())


def test_create_job_uses_manual_time_only_when_ocr_fails(
    tmp_path: Path, monkeypatch
) -> None:
    attempted_offsets: list[float] = []

    async def fake_probe(_self, _path):
        return MediaProbe(
            duration=120.0,
            format_name="mpegts",
            video_codec="h264",
            audio_codec="aac",
        )

    async def failed_time_reference(_self, _source, _frame_path, **kwargs):
        attempted_offsets.append(kwargs["frame_offset_seconds"])
        raise MediaError("timestamp not found")

    async def no_start(_self):
        return None

    async def no_stop(_self):
        return None

    monkeypatch.setattr("slice_helper.media.MediaService.probe", fake_probe)
    monkeypatch.setattr(
        "slice_helper.media.MediaService.detect_time_reference", failed_time_reference
    )
    monkeypatch.setattr(Orchestrator, "start", no_start)
    monkeypatch.setattr(Orchestrator, "stop", no_stop)

    source = tmp_path / "source.ts"
    source.write_bytes(b"source")
    with TestClient(create_app(make_settings(tmp_path))) as client:
        channel_id = create_channel(client)
        response = client.post(
            "/api/jobs",
            json={
                "sourcePath": str(source.resolve()),
                "channelId": channel_id,
                "broadcastDate": "2026-06-20",
                "programStartTime": "2026-06-20T12:30:00+08:00",
            },
        )

    assert response.status_code == 201
    job = response.json()
    assert job["program_start_time"] == "2026-06-20T12:30:00+08:00"
    assert job["time_reference_source"] == "manual_fallback"
    assert attempted_offsets == [0.0, 60.0, 120.0, 180.0, 240.0, 300.0]
    assert job["time_reference_error"].startswith("0s: timestamp not found")
    assert job["time_reference_error"].endswith("300s: timestamp not found")
    assert "fallback" in job["warnings"][0].lower()


def test_create_job_retries_ocr_and_back_calculates_first_frame_time(
    tmp_path: Path, monkeypatch
) -> None:
    attempted_offsets: list[float] = []

    async def fake_probe(_self, _path):
        return MediaProbe(
            duration=600.0,
            format_name="mpegts",
            video_codec="h264",
            audio_codec="aac",
        )

    async def retrying_time_reference(_self, _source, frame_path, **kwargs):
        offset = kwargs["frame_offset_seconds"]
        attempted_offsets.append(offset)
        if offset < 120:
            raise MediaError("timestamp not found")
        frame_path.parent.mkdir(parents=True, exist_ok=True)
        frame_path.write_bytes(b"png")
        observed = datetime.fromisoformat("2026-06-20T12:32:00")
        return TimeReference(
            source_start_time=observed - timedelta(seconds=offset),
            observed_time=observed,
            frame_offset_seconds=offset,
            matched_text="2026-06-20 12:32:00",
            confidence=0.98,
            ocr_texts=(OcrText("2026-06-20 12:32:00", 0.98),),
        )

    async def no_start(_self):
        return None

    async def no_stop(_self):
        return None

    monkeypatch.setattr("slice_helper.media.MediaService.probe", fake_probe)
    monkeypatch.setattr(
        "slice_helper.media.MediaService.detect_time_reference",
        retrying_time_reference,
    )
    monkeypatch.setattr(Orchestrator, "start", no_start)
    monkeypatch.setattr(Orchestrator, "stop", no_stop)

    source = tmp_path / "source.ts"
    source.write_bytes(b"source")
    with TestClient(create_app(make_settings(tmp_path))) as client:
        channel_id = create_channel(client)
        response = client.post(
            "/api/jobs",
            json={
                "sourcePath": str(source.resolve()),
                "channelId": channel_id,
                "broadcastDate": "2026-06-20",
            },
        )

    assert response.status_code == 201
    job = response.json()
    assert attempted_offsets == [0.0, 60.0, 120.0]
    assert job["program_start_time"] == "2026-06-20T12:30:00"
    assert job["time_reference_frame_offset"] == 120.0
    assert job["status"] == "pending_schedule"
    assert "120s" in job["warnings"][0]


def test_missing_time_stops_until_page_correction_and_resyncs_real_times(
    tmp_path: Path, monkeypatch
) -> None:
    async def fake_probe(_self, _path):
        return MediaProbe(
            duration=120.0,
            format_name="mpegts",
            video_codec="h264",
            audio_codec="aac",
        )

    async def failed_time_reference(_self, _source, _frame_path, **_kwargs):
        raise MediaError("timestamp not found")

    async def no_start(_self):
        return None

    async def no_stop(_self):
        return None

    monkeypatch.setattr("slice_helper.media.MediaService.probe", fake_probe)
    monkeypatch.setattr(
        "slice_helper.media.MediaService.detect_time_reference", failed_time_reference
    )
    monkeypatch.setattr(Orchestrator, "start", no_start)
    monkeypatch.setattr(Orchestrator, "stop", no_stop)

    source = tmp_path / "source.ts"
    source.write_bytes(b"source")
    configured = make_settings(tmp_path)
    with TestClient(create_app(configured)) as client:
        channel_id = create_channel(client)
        created = client.post(
            "/api/jobs",
            json={
                "sourcePath": str(source.resolve()),
                "channelId": channel_id,
                "broadcastDate": "2026-06-20",
            },
        )
        assert created.status_code == 201
        job = created.json()
        job_id = job["id"]
        assert job["status"] == "stopped"
        assert job["program_start_time"] is None
        assert "job stopped" in job["error_message"]

        async def add_existing_result() -> None:
            database = Database(configured.database_path)
            window = await database.upsert_window(job_id, 0, 60.0, 120.0)
            attempt = await database.create_attempt(window["id"], 1, "task-time-fix")
            await database.update_attempt(attempt["id"], status="completed")
            await database.replace_window_segments(
                job_id,
                window["id"],
                [
                    {
                        "source_index": 0,
                        "accepted": 1,
                        "reason": "",
                        "local_start": 10.0,
                        "local_end": 20.0,
                        "global_start": 70.0,
                        "global_end": 80.0,
                        "absolute_start": None,
                        "absolute_end": None,
                        "title": "测试片段",
                        "content_type": "新闻",
                        "news_event_type": "",
                        "topic": "",
                        "keywords_json": "[]",
                        "summary": "",
                        "segment_url": "",
                        "cover_img_url": "",
                        "raw_json": "{}",
                    }
                ],
            )

        asyncio.run(add_existing_result())
        corrected = client.patch(
            f"/api/jobs/{job_id}/time-reference",
            json={"programStartTime": "2026-06-20T12:00:00+08:00"},
        )
        assert corrected.status_code == 200
        payload = corrected.json()
        assert payload["updatedSegmentCount"] == 1
        assert payload["job"]["status"] == "paused"
        assert payload["job"]["time_reference_source"] == "manual_override"

        detail = client.get(f"/api/jobs/{job_id}").json()
        assert detail["attempts"][0]["program_start_time"] == (
            "2026-06-20T12:01:00+08:00"
        )
        segment = client.get(f"/api/jobs/{job_id}/segments").json()[0]
        assert segment["absolute_start"] == "2026-06-20T12:01:10+08:00"
        assert segment["absolute_end"] == "2026-06-20T12:01:20+08:00"

        resumed = client.post(f"/api/jobs/{job_id}/resume")
        assert resumed.status_code == 200
        assert resumed.json()["status"] == "pending_schedule"


def test_channel_date_overwrite_pagination_and_excel_export(
    tmp_path: Path, monkeypatch
) -> None:
    async def fake_probe(_self, _path):
        return MediaProbe(
            duration=120.0,
            format_name="mpegts",
            video_codec="h264",
            audio_codec="aac",
        )

    async def failed_time_reference(_self, _source, _frame_path, **_kwargs):
        raise MediaError("timestamp not found")

    async def no_start(_self):
        return None

    async def no_stop(_self):
        return None

    monkeypatch.setattr("slice_helper.media.MediaService.probe", fake_probe)
    monkeypatch.setattr(
        "slice_helper.media.MediaService.detect_time_reference", failed_time_reference
    )
    monkeypatch.setattr(Orchestrator, "start", no_start)
    monkeypatch.setattr(Orchestrator, "stop", no_stop)

    first_source = tmp_path / "first.ts"
    second_source = tmp_path / "second.ts"
    first_source.write_bytes(b"first")
    second_source.write_bytes(b"second")
    configured = make_settings(tmp_path)

    with TestClient(create_app(configured)) as client:
        channel_id = create_channel(client, "测试频道")
        duplicate_channel = client.post(
            "/api/channels", json={"name": "  测试频道  "}
        )
        assert duplicate_channel.status_code == 409
        empty_channel_id = create_channel(client, "待删除频道")
        assert client.delete(f"/api/channels/{empty_channel_id}").status_code == 204

        common = {
            "channelId": channel_id,
            "broadcastDate": "2026-06-20",
            "programStartTime": "2026-06-20T00:00:00+08:00",
        }
        first = client.post(
            "/api/jobs", json={**common, "sourcePath": str(first_source.resolve())}
        )
        assert first.status_code == 201
        first_id = first.json()["id"]

        conflict = client.post(
            "/api/jobs", json={**common, "sourcePath": str(second_source.resolve())}
        )
        assert conflict.status_code == 409
        assert conflict.json()["detail"]["code"] == "channel_date_exists"

        database = Database(configured.database_path)
        asyncio.run(database.update_job(first_id, status="completed"))
        replacement = client.post(
            "/api/jobs",
            json={
                **common,
                "sourcePath": str(second_source.resolve()),
                "overwrite": True,
            },
        )
        assert replacement.status_code == 201
        replacement_id = replacement.json()["id"]
        historical = asyncio.run(database.get_job(first_id))
        assert historical["superseded_by_job_id"] == replacement_id
        assert historical["superseded_at"]

        active_conflict = client.post(
            "/api/jobs",
            json={
                **common,
                "sourcePath": str(first_source.resolve()),
                "overwrite": True,
            },
        )
        assert active_conflict.status_code == 409
        assert active_conflict.json()["detail"]["code"] == "channel_date_active"

        async def add_results() -> None:
            await database.update_job(replacement_id, status="completed")
            window = await database.upsert_window(replacement_id, 0, 0, 120)
            attempt = await database.create_attempt(window["id"], 1, "task-export-1")
            await database.update_attempt(
                attempt["id"], status="completed", service_status="completed", progress=100
            )
            base = {
                "reason": "",
                "local_start": 0.0,
                "local_end": 30.0,
                "global_start": 0.0,
                "global_end": 30.0,
                "absolute_start": "2026-06-20T00:00:00+08:00",
                "absolute_end": "2026-06-20T00:00:30+08:00",
                "content_type": "新闻",
                "news_event_type": "时政要闻",
                "topic": "时政",
                "keywords_json": '["关键词一", "关键词二"]',
                "segment_url": "http://media.test/accepted.mp4",
                "cover_img_url": "http://media.test/accepted.jpg",
            }
            await database.replace_window_segments(
                replacement_id,
                window["id"],
                [
                    {
                        **base,
                        "source_index": 0,
                        "accepted": 1,
                        "title": "最终采用标题",
                        "summary": "界面未显示的完整摘要",
                        "raw_json": '{"title":"任务原标题","contentType":"广告"}',
                    },
                    {
                        **base,
                        "source_index": 1,
                        "accepted": 0,
                        "reason": "handoff",
                        "title": "不应导出的舍弃标题",
                        "summary": "舍弃摘要",
                        "raw_json": "{}",
                    },
                ],
            )

        asyncio.run(add_results())
        accepted_segment = client.get(
            f"/api/jobs/{replacement_id}/segments", params={"acceptedOnly": True}
        ).json()[0]
        assert accepted_segment["ignored"] is False
        edited = client.patch(
            f"/api/jobs/{replacement_id}/segments/{accepted_segment['id']}",
            json={
                "title": "手工修改标题",
                "contentType": "纪录片",
                "ignored": True,
            },
        )
        assert edited.status_code == 200
        assert edited.json()["title"] == "手工修改标题"
        assert edited.json()["content_type"] == "纪录片"
        assert edited.json()["ignored"] is True
        task_values = client.get(
            f"/api/jobs/{replacement_id}/segments/{accepted_segment['id']}/task-values"
        )
        assert task_values.status_code == 200
        assert task_values.json() == {
            "title": "任务原标题",
            "contentType": "广告",
        }
        assert client.patch(
            f"/api/jobs/{replacement_id}/segments/{accepted_segment['id']}",
            json={"contentType": "广告"},
        ).status_code == 422
        restored_from_task = client.patch(
            f"/api/jobs/{replacement_id}/segments/{accepted_segment['id']}",
            json={
                "title": task_values.json()["title"],
                "contentType": task_values.json()["contentType"],
                "restoredFromTask": True,
                "ignored": True,
            },
        )
        assert restored_from_task.status_code == 200
        assert restored_from_task.json()["title"] == "任务原标题"
        assert restored_from_task.json()["content_type"] == "广告"
        edited = client.patch(
            f"/api/jobs/{replacement_id}/segments/{accepted_segment['id']}",
            json={"title": "手工修改标题", "contentType": "纪录片", "ignored": True},
        )
        assert edited.status_code == 200
        assert client.patch(
            f"/api/jobs/{replacement_id}/segments/{accepted_segment['id']}",
            json={"contentType": "专题"},
        ).status_code == 422
        assert client.patch(
            f"/api/jobs/{replacement_id}/segments/{accepted_segment['id']}", json={}
        ).status_code == 422
        assert client.patch(
            f"/api/jobs/{replacement_id}/segments/999999", json={"ignored": True}
        ).status_code == 404
        assert client.get(
            f"/api/jobs/{replacement_id}/segments/999999/task-values"
        ).status_code == 404
        manifest = client.get(f"/api/jobs/{replacement_id}/result").json()
        manifest_segment = next(
            item for item in manifest["segments"] if item["id"] == accepted_segment["id"]
        )
        assert manifest["schemaVersion"] == 6
        assert manifest_segment["title"] == "手工修改标题"
        assert manifest_segment["content_type"] == "纪录片"
        assert manifest_segment["ignored"] is True

        second_date = client.post(
            "/api/jobs",
            json={
                "sourcePath": str(first_source.resolve()),
                "channelId": channel_id,
                "broadcastDate": "2026-06-21",
            },
        )
        assert second_date.status_code == 201

        page_one = client.get("/api/jobs", params={"page": 1, "pageSize": 1}).json()
        page_two = client.get("/api/jobs", params={"page": 2, "pageSize": 1}).json()
        assert page_one["total"] == 2
        assert page_one["totalPages"] == 2
        assert page_one["items"][0]["id"] != page_two["items"][0]["id"]
        filtered = client.get(
            "/api/jobs",
            params={"channelId": channel_id, "broadcastDate": "2026-06-20"},
        ).json()
        assert filtered["total"] == 1
        assert filtered["items"][0]["id"] == replacement_id

        ignored_export = client.get(f"/api/channels/{channel_id}/export.xlsx")
        ignored_workbook = openpyxl.load_workbook(io.BytesIO(ignored_export.content))
        ignored_values = [
            cell.value
            for sheet in ignored_workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
        ]
        assert "手工修改标题" not in ignored_values
        assert "界面未显示的完整摘要" not in ignored_values

        restored = client.patch(
            f"/api/jobs/{replacement_id}/segments/{accepted_segment['id']}",
            json={"ignored": False},
        )
        assert restored.status_code == 200
        assert restored.json()["ignored"] is False

        exported = client.get(f"/api/channels/{channel_id}/export.xlsx")
        assert exported.status_code == 200
        assert exported.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        exported_workbook = openpyxl.load_workbook(io.BytesIO(exported.content))
        assert exported_workbook.sheetnames == ["2026-06-20", "2026-06-21"]
        first_sheet = exported_workbook["2026-06-20"]
        assert tuple(cell.value for cell in first_sheet[1][:42]) == HEADERS
        assert first_sheet["B2"].value == "手工修改标题"
        assert first_sheet["C2"].value == "关键词一, 关键词二"
        assert first_sheet["D2"].value == "界面未显示的完整摘要"
        assert first_sheet["E2"].value == "测试频道"
        assert first_sheet["F2"].value == datetime(2026, 6, 20, 0, 0, 0)
        assert first_sheet["G2"].value == datetime(2026, 6, 20, 0, 0, 30)
        assert first_sheet["H2"].value == datetime(2026, 6, 20, 0, 0, 0)
        assert first_sheet["I2"].value == "纪录片"
        assert first_sheet["J2"].value == "时政要闻"
        assert first_sheet["T2"].value is None
        assert first_sheet["V2"].value is None
        exported_values = [cell.value for row in first_sheet.iter_rows() for cell in row]
        assert "task-export-1" not in exported_values
        assert "http://media.test/accepted.mp4" not in exported_values
        assert "时政" not in exported_values
        assert "不应导出的舍弃标题" not in exported_values
        assert first_sheet.auto_filter.ref == "A1:AP2"
        assert all(cell.data_type != "f" for row in first_sheet.iter_rows() for cell in row)

        style_template = openpyxl.load_workbook(TEMPLATE_PATH).active
        assert first_sheet.freeze_panes == style_template.freeze_panes == "O2"
        assert first_sheet.row_dimensions[1].height == style_template.row_dimensions[1].height == 45
        for column in ("A", "B", "F", "H", "U", "V", "AP"):
            assert (
                first_sheet.column_dimensions[column].width
                == style_template.column_dimensions[column].width
            )
        for column in range(1, 43):
            assert first_sheet.cell(1, column)._style == style_template.cell(1, column)._style
            assert first_sheet.cell(2, column)._style == style_template.cell(2, column)._style
            assert (
                exported_workbook["2026-06-21"].cell(1, column)._style
                == style_template.cell(1, column)._style
            )
            assert (
                exported_workbook["2026-06-21"].cell(2, column)._style
                == style_template.cell(2, column)._style
            )

        app_js = client.get("/static/app.js").text
        assert '["类型2", "新闻事件类型", segment.news_event_type || ""]' in app_js
        assert '["标签", unmapped, ""]' in app_js
        assert '["视频链接"' not in app_js

        renamed = client.patch(
            f"/api/channels/{channel_id}", json={"name": "测试频道（新）"}
        )
        assert renamed.status_code == 200
        assert renamed.json()["name"] == "测试频道（新）"
        assert client.delete(f"/api/channels/{channel_id}").status_code == 409


def test_manual_segment_merge_api_manifest_and_excel(tmp_path: Path, monkeypatch) -> None:
    async def no_start(_self):
        return None

    async def no_stop(_self):
        return None

    monkeypatch.setattr(Orchestrator, "start", no_start)
    monkeypatch.setattr(Orchestrator, "stop", no_stop)
    configured = make_settings(tmp_path)

    with TestClient(create_app(configured)) as client:
        database = Database(configured.database_path)

        async def add_results() -> tuple[str, str]:
            channel = await database.create_channel("接口合并频道")
            job_id = "merge-api-job"
            await database.create_job(
                {
                    "id": job_id,
                    "source_path": str(tmp_path / "source.ts"),
                    "source_size": 10,
                    "source_mtime_ns": 20,
                    "source_duration": 60.0,
                    "template_id": "general",
                    "language": "zh",
                    "channel_id": channel["id"],
                    "channel_name": channel["name"],
                    "broadcast_date": "2026-08-11",
                    "program_start_time": "2026-08-11T08:00:00+08:00",
                    "cut_mode": "copy",
                    "total_windows": 1,
                    "status": "completed",
                }
            )
            window = await database.upsert_window(job_id, 0, 0, 60)
            await database.replace_window_segments(
                job_id,
                window["id"],
                [
                    {
                        "source_index": index,
                        "accepted": 1,
                        "reason": "",
                        "local_start": index * 10.0,
                        "local_end": (index + 1) * 10.0,
                        "global_start": index * 10.0,
                        "global_end": (index + 1) * 10.0,
                        "title": title,
                        "content_type": "新闻",
                        "news_event_type": "时政要闻",
                        "topic": "主题",
                        "keywords_json": '["关键词"]',
                        "summary": f"{title}摘要",
                        "segment_url": "",
                        "cover_img_url": "",
                        "raw_json": "{}",
                    }
                    for index, title in enumerate(("第一条", "第二条", "第三条"))
                ],
            )
            return job_id, str(channel["id"])

        job_id, channel_id = asyncio.run(add_results())
        segments = client.get(f"/api/jobs/{job_id}/segments").json()
        first, primary = segments[:2]
        preview_response = client.post(
            f"/api/jobs/{job_id}/segment-merges/preview",
            json={
                "segmentIds": [first["id"], primary["id"]],
                "primarySegmentId": primary["id"],
            },
        )
        assert preview_response.status_code == 200
        preview = preview_response.json()
        assert preview["result"]["title"] == "第二条"
        assert preview["result"]["globalStart"] == 0
        assert preview["result"]["globalEnd"] == 20

        created = client.post(
            f"/api/jobs/{job_id}/segment-merges",
            json={
                "segmentIds": preview["segmentIds"],
                "primarySegmentId": preview["primarySegmentId"],
                "previewToken": preview["previewToken"],
            },
        )
        assert created.status_code == 201
        merge_id = created.json()["id"]
        merged_segments = client.get(f"/api/jobs/{job_id}/segments").json()
        merge_result = next(item for item in merged_segments if item["record_kind"] == "merge")
        assert merge_result["manual_merge"] is True
        assert merge_result["member_count"] == 2
        assert len([item for item in merged_segments if item.get("active_merge_id")]) == 2

        locked_member = client.patch(
            f"/api/jobs/{job_id}/segments/{first['id']}", json={"title": "不应保存"}
        )
        assert locked_member.status_code == 409
        edited = client.patch(
            f"/api/jobs/{job_id}/segment-merges/{merge_id}",
            json={"title": "合并后的标题", "contentType": "纪录片"},
        )
        assert edited.status_code == 200
        assert edited.json()["title"] == "合并后的标题"
        assert edited.json()["record_kind"] == "merge"

        manifest = client.get(f"/api/jobs/{job_id}/result").json()
        assert manifest["schemaVersion"] == 6
        assert manifest["manualMerges"][0]["status"] == "active"
        assert any(item["record_kind"] == "merge" for item in manifest["segments"])

        exported = client.get(f"/api/channels/{channel_id}/export.xlsx")
        workbook = openpyxl.load_workbook(io.BytesIO(exported.content))
        sheet = workbook["2026-08-11"]
        assert sheet.max_row == 3
        assert sheet["B2"].value == "合并后的标题"
        assert sheet["AM2"].value == "是"
        exported_titles = {sheet["B2"].value, sheet["B3"].value}
        assert exported_titles == {"合并后的标题", "第三条"}

        cancelled = client.delete(f"/api/jobs/{job_id}/segment-merges/{merge_id}")
        assert cancelled.status_code == 204
        restored = client.get(f"/api/jobs/{job_id}/segments").json()
        assert not any(item["record_kind"] == "merge" for item in restored)
        assert not any(item.get("active_merge_id") for item in restored)
