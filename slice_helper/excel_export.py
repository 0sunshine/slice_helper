from __future__ import annotations

import io
import json
import re
from collections import defaultdict
from copy import copy
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


CHINA_TIMEZONE = timezone(timedelta(hours=8))
TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "deep_epg_template.xlsx"


# Keep this order identical to the supplied 深度 EPG workbook template.
HEADERS = (
    "节目ID",
    "program_name",
    "关键字",
    "摘要",
    "channel_name",
    "begin_time",
    "end_time",
    "日期",
    "类型1",
    "类型2",
    "节目是否首播",
    "内容是否首播",
    "是否栏目",
    "栏目名称",
    "节目",
    "集数",
    "是否黄金时段",
    "是否国产",
    "是否收官",
    "标签",
    "是否含广告",
    "处理方式",
    "改后节目名称",
    "改后开始时间",
    "改后结束时间",
    "改后日期",
    "改后类型1",
    "改后类型2",
    "改后节目是否首播",
    "改后内容是否首播",
    "改后是否栏目",
    "改后栏目名称",
    "改后节目",
    "改后集数",
    "改后是否黄金时段",
    "改后是否国产",
    "改后是否收官",
    "改后标签",
    "是否合并",
    "是否拆条",
    "改后节目曾用名",
    "是否立即",
)


def safe_export_filename(channel_name: str) -> str:
    safe = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", channel_name).strip(" ._")
    return f"{safe or '频道'}-深度EPG业务.xlsx"


def build_channel_workbook(export: dict) -> bytes:
    channel = export["channel"]
    segments_by_date: dict[str, list[dict]] = defaultdict(list)
    for segment in export["segments"]:
        segments_by_date[str(segment["broadcast_date"])].append(segment)
    job_dates = [str(job["broadcast_date"]) for job in export["jobs"]]
    sheet_dates = sorted(dict.fromkeys(job_dates))

    workbook = openpyxl.load_workbook(TEMPLATE_PATH)
    template_sheet = workbook.worksheets[0]
    workbook.properties.title = f"{channel['name']} 深度EPG业务"
    workbook.properties.subject = "按业务日期汇总的最终采用拆条结果"
    workbook.properties.company = "iSlice"

    if not sheet_dates:
        template_sheet.title = "暂无数据"
    else:
        sheets = [template_sheet]
        for _ in sheet_dates[1:]:
            sheets.append(workbook.copy_worksheet(template_sheet))
        for sheet, broadcast_date in zip(sheets, sheet_dates, strict=True):
            sheet.title = broadcast_date
            _write_date_sheet(
                sheet,
                channel_name=str(channel["name"]),
                broadcast_date=broadcast_date,
                segments=segments_by_date.get(broadcast_date, []),
            )

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_date_sheet(
    sheet: Worksheet,
    *,
    channel_name: str,
    broadcast_date: str,
    segments: list[dict],
) -> None:
    business_date = date.fromisoformat(broadcast_date)
    for index, segment in enumerate(segments, start=2):
        if index > 2:
            _copy_template_row(sheet, index)
        absolute_start = _excel_datetime(segment.get("absolute_start"))
        absolute_end = _excel_datetime(segment.get("absolute_end"))
        export_date = absolute_start.date() if absolute_start else business_date
        keywords = _keywords(segment.get("keywords_json"))

        # Only these template columns have an agreed mapping. All other cells keep
        # the blank value and exact style copied from the supplied workbook.
        sheet.cell(index, 2).value = str(segment.get("title") or "")
        sheet.cell(index, 3).value = ", ".join(keywords)
        sheet.cell(index, 4).value = str(segment.get("summary") or "")
        sheet.cell(index, 5).value = channel_name
        sheet.cell(index, 6).value = absolute_start
        sheet.cell(index, 7).value = absolute_end
        sheet.cell(index, 8).value = export_date
        sheet.cell(index, 9).value = str(segment.get("content_type") or "")
        sheet.cell(index, 10).value = str(segment.get("news_event_type") or "")

    sheet.auto_filter.ref = f"A1:AP{max(2, len(segments) + 1)}"


def _copy_template_row(sheet: Worksheet, target_row: int) -> None:
    source_row = 2
    source_dimension = sheet.row_dimensions[source_row]
    target_dimension = sheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel
    target_dimension.collapsed = source_dimension.collapsed

    for column in range(1, len(HEADERS) + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        target._style = copy(source._style)
        if source.has_style:
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)


def _excel_datetime(value: object) -> datetime | None:
    if not value:
        return None
    parsed = datetime.fromisoformat(str(value))
    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(CHINA_TIMEZONE).replace(tzinfo=None)
    return parsed


def _keywords(value: object) -> list[str]:
    try:
        parsed = json.loads(str(value or "[]"))
    except (TypeError, ValueError, json.JSONDecodeError):
        return []
    return [str(item) for item in parsed] if isinstance(parsed, list) else []
